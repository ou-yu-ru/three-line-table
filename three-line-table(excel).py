import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_three_line_table(dataframe):
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 将DataFrame写入工作表
    for r in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(r)

    # 添加三线
    for col in range(1, len(dataframe.columns) + 1):
        ws.cell(row=1, column=col).border = Border(bottom=Side(style='thin'), top=Side(style='thick'))
        ws.cell(row=len(dataframe)+1, column=col).border = Border(bottom=Side(style='thick'))

    # 将工作簿保存到BytesIO对象中
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # 将指针移回起始位置
    return output

# Streamlit应用程序
st.title('Excel 三线表生成器')
st.write("👀注意：表格从左上角A1开始。")
uploaded_file = st.file_uploader("上传一个Excel文件", type=["xlsx"])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    st.write("数据预览:")
    st.dataframe(df)

    if st.button("生成三线表"):
        output = create_three_line_table(df)
        st.download_button(
            label="下载生成的Excel文件",
            data=output,
            file_name="三线表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
